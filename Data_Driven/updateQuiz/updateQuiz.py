# Generated by Selenium IDE
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

from selenium.common.exceptions import NoSuchElementException

class FileExcelReader:
    file = ""
    sheetName = ""

    def __init__(self, file, sheetName):
        self.file = file
        self.sheetName = sheetName

    def getRowCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_row)

    def getColumnCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_column)

    def readData(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestUpdateQuiz():
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.vars = {}
    
    def teardown_method(self):
        self.driver.quit()

    def first_step(self):
        self.driver.get('https://sso.hcmut.edu.vn/cas/login?service=https%3A%2F%2Fe-learning.hcmut.edu.vn%2Flogin%2Findex.php%3FauthCAS%3DCAS')

        self.driver.find_element(By.NAME,"username").send_keys("010354")
        self.driver.find_element(By.NAME,"password").send_keys("010354")
        self.driver.find_element(By.NAME,"submit").click()
        time.sleep(2)

        self.driver.get('https://e-learning.hcmut.edu.vn/course/modedit.php?update=73403&return=1')
        time.sleep(1)

        self.driver.find_element(By.ID, 'collapseElement-1').click()
        time.sleep(5)

    def test_updateQuiz(self, name, open, close, limit, expectedResult):
        self.first_step()
        
        self.driver.execute_script(f"arguments[0].setAttribute('value',{name})", self.driver.find_element(By.ID, 'id_name'))
        if (open != ""):
            self.driver.find_element(By.ID, 'id_timeopen_enabled').click()

        if (close != ""):
            self.driver.find_element(By.ID, 'id_timeclose_enabled').click()

        if (limit != ""):
            self.driver.find_element(By.ID, 'id_timelimit_enabled').click()

        if (open == '1' and close == '0'):
            self.driver.find_element(By.ID, 'id_timeopen_enabled').click()
            self.driver.find_element(By.XPATH, '//*[@id="id_timeclose_year"]/option[text()="2000"]').click()
        
        elif (limit == '-1'):
            self.driver.execute_script("arguments[0].setAttribute('value','-1')", self.driver.find_element(By.ID, 'id_timelimit_number'))
        
        self.driver.find_element(By.ID,"id_submitbutton2").click()
        time.sleep(5)

        if expectedResult == "Success":
            self.driver.find_element(By.ID,'topofscroll')

        elif expectedResult == "MissingValue":
            self.driver.find_element(By.ID,'id_error_name')

        else:
            self.driver.find_element(By.ID,'id_error_timeclose')

        time.sleep(1)
        self.driver.get('https://e-learning.hcmut.edu.vn/login/logout.php?sesskey=ZQQ9wh96tV')
        time.sleep(1)

if __name__ == "__main__":
    excel = FileExcelReader('SecB_updatequiz_data.xlsx', 'Sheet1')
    test = TestUpdateQuiz()
    test.setup_method()
    nRows = excel.getRowCount()
    for row in range(2, nRows + 1):
        name = excel.readData(row,1)
        open = excel.readData(row,2)
        close = excel.readData(row,3)
        limit = excel.readData(row,4)
        expectedResult = excel.readData(row,5)
       
        if name is None:
            name = ""
        if open is None:
            open = ""
        if close is None:
            close = ""
        if limit is None:
            limit = ""

        try:
            result = test.test_updateQuiz(name, open, close, limit, expectedResult)
            excel.writeData("Passed",row,5)
        except:
            excel.writeData("Failed",row,5)

    test.teardown_method()